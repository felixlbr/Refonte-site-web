import sqlite3
from unittest import result
from sqlalchemy import create_engine
import pandas as pd
import numpy as np
import openpyxl
import matplotlib.pyplot as plt

engine = create_engine('sqlite://', echo=False)

#creation des 2 tables
df = pd.read_excel('Sondage.xlsx') 
df.to_sql('sondage', engine, if_exists='replace', index=False)
df = pd.read_excel('Aliments.xlsx')
df.to_sql('aliments', engine, if_exists='append', index=False)

# Requete 1
# ----------------------------------------------------------------


#vérifie si un élement <tableau> existe dans un tableau <tableau>
def existIn(elementT, listT):
  for i in listT:
    if elementT[0] == i[0]:
      return True
  return False


#récupère la liste des personnes pour 1 colonne
def getPers(list, colonne, nom):
  requete = engine.execute(
    "select 'Nom :' || s.Nom || ' Prénom :' || s.Prénom || ' Naissance :' || s.Naissance || ' Adresse :' || s.Adresse || ' Code Postal :' || s.[Code Postal] || ' Ville :' || s.Ville || ' Tél :' || s.Tel From Sondage s INNER JOIN Aliments a ON s." + colonne +
    "=a.alim_code where LOWER(a.alim_ssssgrp_nom_fr) like LOWER('%" + nom +
    "%') OR LOWER(a.alim_nom_fr) like LOWER('%" + nom + "%')")
  final = pd.DataFrame(requete)
  for k in final.values:
    if not existIn(k, list):
      list.append(k.tolist())
  return list


#affiche la liste des personnes pour les 10 aliments
def getListe(liste, produit):
  t = "Personnes qui consomme " + produit
  for i in range(10):
    liste = getPers(liste, "Aliment" + str(i + 1), produit)
  return t, liste


#affichage des personnes en fonction des catégories
listHalal = []
listBio = []
listCasher = []
listVegan = []

def writePers(t1, t2, t3, t4, col1, col2, col3, col4):
  my_wb = openpyxl.Workbook()
  my_sheet = my_wb.active
  i = 0
  j = 0
  k = 0
  l = 0
  my_sheet.cell(row=1, column=1).value = t1
  while i < len(col1):
    my_sheet.cell(row=i + 2, column=1).value = col1[i][0]
    i += 1

  my_sheet.cell(row=1, column=2).value = t2
  while j < len(col2):
    my_sheet.cell(row=j + 2, column=2).value = col2[j][0]
    j += 1

  my_sheet.cell(row=1, column=3).value = t3
  while k < len(col3):
    my_sheet.cell(row=k + 2, column=3).value = col3[k][0]
    k += 1

  my_sheet.cell(row=1, column=4).value = t4
  while l < len(col4):
    my_sheet.cell(row=l + 2, column=4).value = col4[l][0]
    l += 1

  my_wb.save("ListePers.xlsx")


t1, listHalal = getListe(listHalal, "halal")
t2, listBio = getListe(listBio, "bio")
t3, listCasher = getListe(listCasher, "casher")
t4, listVegan = getListe(listVegan, "vegan")
writePers(t1, t2, t3, t4, listHalal, listBio, listCasher, listVegan)

def getTotal():
  requete = engine.execute(
  "select 'Nom :' || s.Nom || ' Prénom :' || s.Prénom || ' Naissance :' || s.Naissance || ' Adresse :' || s.Adresse || ' Code Postal :' || s.[Code Postal] || ' Ville :' || s.Ville || ' Tél :' || s.Tel From Sondage s")
  final = pd.DataFrame(requete)
  return final.values

listReste = getTotal()
listReste = listReste.tolist()

def getReste(listReste, list):
  for i in list:
    if i in listReste:
      listReste.remove(i)
  return listReste

listReste = getReste(listReste, listHalal)
listReste = getReste(listReste, listBio)
listReste = getReste(listReste, listCasher)
listReste = getReste(listReste, listVegan)
# Graphique
#-----------------------------------------------------------------------

x = [len(listReste), len(listHalal), len(listBio), len(listCasher), len(listVegan)]
plt.pie(x, labels = ['Autre', 'Halal', 'Bio', 'Casher', 'Vegan'], colors = ['darkgreen', 'green', 'seagreen', 'mediumseagreen', 'darkseagreen'], autopct = lambda x: str(round(x, 2)) + '%', normalize = True)
plt.title("Part de la population qui consomme :")
plt.savefig("camembert",dpi=500, facecolor='w',orientation='portrait',format=None,transparent=True,bbox_inches=None, pad_inches=0.1,metadata=None)
plt.show()

# Requete 2
# ----------------------------------------------------------------


def getCateg(list, colonne, categorie):
  requete = engine.execute("select a." + categorie +
                           " From Sondage s INNER JOIN Aliments a ON s." +
                           colonne + "=a.alim_code")
  final = pd.DataFrame(requete)
  for k in final.values:
    list.append(k[0])
  return (list)


def maxCount(list):
  return {k: list.count(k) for k in list}


listeCateg1 = []
listeCateg2 = []
listeCateg3 = []


def affichageCateg(liste, categorie):
  for i in range(10):
    liste = getCateg(liste, "Aliment" + str(i + 1), categorie)
  return maxCount(liste)


def getTableau(dico):
  col1 = []
  col2 = []
  t1 = "Categorie"
  t2 = "Nombre de vote"
  for cle, valeur in dico.items():
    col1.append(cle)
    col2.append(valeur)
  return t1, col1, t2, col2


def writeCateg(titre, t1, col1, t2, col2):
  data = pd.DataFrame({t1: col1, t2: col2})
  data.to_excel(titre + '.xlsx', sheet_name='sheet1', index=False)


categ1t1, categ1col1, categ1t2, categ1col2 = getTableau(
  affichageCateg(listeCateg1, "alim_grp_nom_fr"))
writeCateg('Categorie1', categ1t1, categ1col1, categ1t2, categ1col2)

categ2t1, categ2col1, categ2t2, categ2col2 = getTableau(
  affichageCateg(listeCateg2, "alim_ssgrp_nom_fr"))
writeCateg('Categorie2', categ2t1, categ2col1, categ2t2, categ2col2)

categ3t1, categ3col1, categ3t2, categ3col2 = getTableau(
  affichageCateg(listeCateg3, "alim_ssssgrp_nom_fr"))
writeCateg('Categorie3', categ3t1, categ3col1, categ3t2, categ3col2)

plt.barh(categ1col1, categ1col2, color = 'darkgreen')
plt.savefig("diagrammeBat1", dpi=500, facecolor='w', bbox_inches='tight')
plt.show()
plt.barh(categ2col1, categ2col2, color = 'green')
plt.savefig("diagrammeBat2",dpi=500, facecolor='w', bbox_inches='tight')
plt.show()
plt.barh(categ3col1, categ3col2, color = 'seagreen')
plt.savefig("diagrammeBat3",dpi=500, facecolor='w', bbox_inches='tight')
plt.show()